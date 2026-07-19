"""
enrich_catalog_vision.py
------------------------
Batch-processes products 1-100:
1. Loads URL, keywords, cost, and inventory from Fashion Apparel.xlsx
2. Finds corresponding local image (e.g. 1.jpg, 2.webp) in frontend/public/images/
3. Calls Gemini 3.5 Flash Vision to generate structured metadata:
   - contextual_description (no hashtags)
   - material, color, nature, age_range, category
4. Updates local_catalog.json and Supabase database.
5. Supports resuming and limiting run counts for dry runs.
"""
import os
import sys
import json
import time
import logging
import openpyxl
from PIL import Image
import google.generativeai as genai
from dotenv import load_dotenv

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger("enrich_catalog")

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
ENV_PATH = os.path.join(BASE_DIR, "backend", ".env")
load_dotenv(ENV_PATH)

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")

LOCAL_CATALOG_FILE = os.path.join(BASE_DIR, "backend", "local_catalog.json")
EXCEL_FILE = os.path.join(BASE_DIR, "Fashion Apparel.xlsx")
IMAGES_DIR = os.path.join(BASE_DIR, "frontend", "public", "images")

# Configure Gemini
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)
else:
    logger.error("GEMINI_API_KEY not found in environment!")

def find_image_file(product_id):
    """Finds image file matching product_id in IMAGES_DIR with any extension."""
    if not os.path.exists(IMAGES_DIR):
        return None
    for fname in os.listdir(IMAGES_DIR):
        name, ext = os.path.splitext(fname)
        if name == str(product_id) and ext.lower() in ['.jpg', '.jpeg', '.png', '.webp', '.avif']:
            return os.path.join(IMAGES_DIR, fname)
    return None

def call_gemini_vision(image_path, keywords):
    """Passes local image and keywords to Gemini 3.5 Flash for visual metadata extraction."""
    try:
        img = Image.open(image_path)
    except Exception as e:
        logger.error(f"Failed to open image {image_path}: {e}")
        return None

    prompt = f"""
    Analyze the apparel item shown in this image. I have also provided some search keywords related to it:
    Keywords: "{keywords}"

    You must analyze both the image and the keywords, and generate a structured JSON object containing:
    1. "contextual_description": A premium, highly detailed 2-3 sentence style description of this fashion item (e.g. detailing fabric fall, necklines, sleeves, style context, and matching events). STRICTLY do NOT include any hashtags.
    2. "material": The dominant fabric/material (choose from: Silk, Cotton, Velvet, Linen, Denim, Wool, Polyester, Rayon, Georgette, Organza, Leather, Synthetic).
    3. "color": The dominant color(s) of the item (choose from: Red, Maroon, Yellow, Gold, White, Black, Blue, Saffron, Pink, Magenta, Green, Grey, Multi).
    4. "nature": The style profile of the item (choose from: everyday, festive, winter, streetwear).
    5. "age_range": The target age group (choose from: Gen-Z, Millennial, Gen-X).
    6. "category": The category tag (choose from: Ethnic, Western, Accessory, Footwear).

    You MUST output ONLY a valid JSON object matching the structure above (no markdown wrapping, no additional text).
    Example:
    {{
      "contextual_description": "An elegant red and yellow Banarasi saree woven in premium cotton silk, featuring delicate gold zari work perfect for evening festive rituals.",
      "material": "Cotton",
      "color": "Red",
      "nature": "festive",
      "age_range": "Millennial",
      "category": "Ethnic"
    }}
    """

    max_retries = 1
    for attempt in range(max_retries):
        try:
            model = genai.GenerativeModel('gemini-3.5-flash')
            response = model.generate_content([img, prompt])
            text = response.text.strip()
            # Clean markdown if present
            if text.startswith("```"):
                text = text.split("\n", 1)[1].rsplit("\n", 1)[0].strip()
                if text.startswith("json"):
                    text = text[4:].strip()
            return json.loads(text)
        except Exception as e:
            logger.warning(f"Gemini API call failed or rate-limited: {e}")
            return None
    return None

def generate_fallback_metadata(keywords):
    kw = str(keywords).lower()
    
    # 1. Determine Category
    category = "casual"
    if any(k in kw for k in ["saree", "kurta", "sherwani", "ethnic", "nehru", "chaniya", "lehenga", "anarkali"]):
        category = "ethnic"
    elif any(k in kw for k in ["gown", "dress", " farewell", " blazer", " tuxedo"]):
        category = "western"
    elif any(k in kw for k in ["jacket", "sweater", "fleece", "hoodie", "cardigan", "muffler", "shawl"]):
        category = "winter"
    elif any(k in kw for k in ["jogger", "hoodie", "sneaker", "streetwear"]):
        category = "streetwear"
    
    # 2. Determine Material
    material = "Cotton"
    for mat in ["silk", "linen", "velvet", "denim", "wool", "georgette", "chiffon", "crepe", "satin", "leather", "fleece", "suede"]:
        if mat in kw:
            material = mat.capitalize()
            break
            
    # 3. Determine Color
    color = "Multi"
    for col in ["red", "yellow", "maroon", "black", "white", "gold", "blue", "green", "pink", "grey", "brown", "beige", "cream", "orange"]:
        if col in kw:
            color = col.capitalize()
            break
            
    # 4. Determine Nature
    nature = "casual"
    if any(k in kw for k in ["puja", "diwali", "chhath", "bihar diwas", "wedding", "sangeet", "festive", "ceremonial", "haldi"]):
        nature = "festive"
    elif any(k in kw for k in ["farewell", "graduation", "formal", "office", "interview"]):
        nature = "formal"
        
    # 5. Age Range
    age_range = "Gen Z"
    if any(k in kw for k in ["wedding", "saree", "formal", "office", "farewell", "classic", "elegant"]):
        age_range = "Millennial"
        
    # 6. Generate Contextual Description (no hashtags)
    desc = f"An elegant {color.lower()} {material.lower()} garment designed beautifully for any occasion. Featuring a premium fit and comfort details."
    if "saree" in kw:
        desc = f"A gorgeous {color.lower()} {material.lower()} saree designed with beautiful traditional details. Drapes elegantly and offers a lightweight comfort, making it a perfect choice for festive celebrations."
    elif "kurta" in kw:
        desc = f"A classic {color.lower()} {material.lower()} kurta set designed for a smart look. Features a premium texture and lightweight weave that is comfortable to wear for celebrations and ceremonies."
    elif "jacket" in kw or "hoodie" in kw or "sweater" in kw:
        desc = f"A stylish and warm {color.lower()} {material.lower()} jacket. Designed with premium insulation and modern fit, making it perfect for winter layering."
    elif "gown" in kw or "dress" in kw:
        desc = f"A sophisticated {color.lower()} {material.lower()} dress featuring a sleek cut and premium drape. Perfect for farewell parties, dinners, or formal evening gatherings."
    else:
        # Custom build description from keyword chunks
        clean_kw = keywords.replace("", "'").replace("'", "")
        desc = f"A premium quality {clean_kw.lower()}. Crafted from fine {material.lower()} fabric, offering both comfort and style, perfectly suited for {nature} styling."

    return {
        "contextual_description": desc,
        "material": material,
        "color": color,
        "nature": nature,
        "age_range": age_range,
        "category": category
    }

def process_batch(limit=None):
    if not os.path.exists(EXCEL_FILE):
        logger.error(f"Excel file not found at {EXCEL_FILE}")
        return

    # Load local_catalog.json
    with open(LOCAL_CATALOG_FILE, "r", encoding="utf-8") as f:
        catalog = json.load(f)
    catalog_map = {p["id"]: p for p in catalog}

    # Open Excel sheet
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active

    logger.info(f"Loaded Excel rows. Total items to check: {ws.max_row}")

    supabase = None
    if SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY:
        try:
            from supabase import create_client
            supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
            logger.info("Connected to Supabase.")
        except Exception as e:
            logger.error(f"Failed to connect to Supabase: {e}")

    processed_count = 0
    # Excel rows: Row 1 matches ID 1, Row 100 matches ID 100
    for r_idx in range(1, min(101, ws.max_row + 1)):
        pid = r_idx
        row = ws[r_idx]
        url = row[0].value
        keywords = row[1].value
        
        try:
            price = int(row[2].value)
        except (ValueError, TypeError):
            price = 999

        try:
            inventory = int(row[3].value)
        except (ValueError, TypeError):
            inventory = 15

        # Check if product is in catalog
        if pid not in catalog_map:
            logger.warning(f"Product ID {pid} not found in local_catalog.json, skipping...")
            continue

        p_data = catalog_map[pid]
        
        # Check if already processed (has material and color) to support resume
        if p_data.get("material") and p_data.get("color") and not limit:
            logger.info(f"Product {pid} already has enriched metadata, skipping.")
            continue

        image_path = find_image_file(pid)
        if not image_path:
            logger.warning(f"Image for Product ID {pid} not found in images directory!")
            continue

        logger.info(f"Processing Product ID {pid} ({keywords})...")
        fallback_used = False
        llm_data = call_gemini_vision(image_path, keywords)

        if not llm_data:
            logger.warning(f"Failed to get LLM response for ID {pid}, using fallback generator...")
            llm_data = generate_fallback_metadata(keywords)
            fallback_used = True

        # Update local catalog item representation
        p_data["description"] = llm_data["contextual_description"]
        p_data["price"] = price
        p_data["inventory"] = inventory
        p_data["material"] = llm_data["material"]
        p_data["color"] = llm_data["color"]
        p_data["nature"] = llm_data["nature"]
        p_data["age_range"] = llm_data["age_range"]
        
        # Category classification
        p_data["category"] = llm_data["category"].lower()

        # Update tags array to feed recommendation engine with structured properties
        new_tags = set(p_data.get("tags", []))
        # Add material, color, nature, age_range, category as matching tags
        new_tags.add(llm_data["material"].lower())
        new_tags.add(llm_data["color"].lower())
        new_tags.add(llm_data["nature"].lower())
        new_tags.add(llm_data["age_range"].lower())
        new_tags.add(llm_data["category"].lower())
        p_data["tags"] = list(new_tags)

        # Sync to Supabase
        if supabase:
            try:
                # Prepare payload
                payload = {
                    "description": p_data["description"],
                    "price": price,
                    "inventory": inventory,
                    "material": p_data["material"],
                    "color": p_data["color"],
                    "nature": p_data["nature"],
                    "age_range": p_data["age_range"],
                    "tags": p_data["tags"],
                    "category": p_data.get("category", "casual")
                }
                supabase.table("products").update(payload).eq("id", pid).execute()
                logger.info(f"  Synced Product ID {pid} to Supabase successfully.")
            except Exception as e:
                logger.error(f"  Failed to sync Product ID {pid} to Supabase: {e}")

        processed_count += 1
        
        # Save local catalog incrementally so we don't lose progress if interrupted
        with open(LOCAL_CATALOG_FILE, "w", encoding="utf-8") as f:
            json.dump(catalog, f, ensure_ascii=False, indent=2)

        # Throttle request rate: only sleep if we actually hit the LLM
        if not fallback_used:
            time.sleep(13.5)
        else:
            time.sleep(0.1)

        if limit and processed_count >= limit:
            logger.info(f"Limit of {limit} reached, stopping.")
            break

    logger.info(f"Vision Enrichment finished. Processed {processed_count} items.")

if __name__ == "__main__":
    limit_val = None
    if len(sys.argv) > 1:
        try:
            limit_val = int(sys.argv[1])
        except ValueError:
            pass
    process_batch(limit=limit_val)
