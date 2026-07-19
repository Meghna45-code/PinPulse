import os
import sys
import json
import time
import logging
import openpyxl
from dotenv import load_dotenv
import google.generativeai as genai
from PIL import Image

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger("enrich_catalog2_vision")

load_dotenv()

# Model config
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)

LOCAL_CATALOG_FILE = os.path.abspath(os.path.join(os.path.dirname(__file__), "local_catalog.json"))
EXCEL_FILE = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "Fashion Apparel2.xlsx"))
IMAGES_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "frontend", "public", "images"))

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")

def find_image_file(pid):
    # Search in public/images for a file named pid.*
    for ext in [".webp", ".png", ".jpg", ".jpeg", ".avif"]:
        p = os.path.join(IMAGES_DIR, f"{pid}{ext}")
        if os.path.exists(p):
            return p
    return None

def call_gemini_vision(image_path, keywords):
    if not GEMINI_API_KEY:
        return None
    try:
        img = Image.open(image_path)
    except Exception as e:
        logger.error(f"Failed to open image {image_path}: {e}")
        return None

    prompt = f"""
    Analyze the fashion apparel product image and the provided metadata keywords: "{keywords}".
    Return a valid JSON object matching this schema:
    {{
      "contextual_description": "A highly detailed, professional, contextual description of the apparel (1-2 sentences), suitable for a premium fashion website. DO NOT include any hashtags.",
      "material": "The dominant fabric material (e.g., Cotton, Silk, Linen, Velvet, Wool, Denim, Polyester)",
      "color": "The primary color or 'Multi' if multi-colored",
      "nature": "The style vibe or occasion suitability: must be one of 'festive', 'casual', 'formal'",
      "age_range": "The target demographic: must be one of 'Gen Z', 'Millennial'",
      "category": "The main category of clothing: must be one of 'ethnic', 'western', 'winter', 'streetwear'"
    }}
    """

    max_retries = 1
    for attempt in range(max_retries):
        try:
            model = genai.GenerativeModel('models/gemini-3.5-flash')
            response = model.generate_content([img, prompt])
            text = response.text.strip()
            if text.startswith("```"):
                text = text.split("\n", 1)[1].rsplit("\n", 1)[0].strip()
                if text.startswith("json"):
                    text = text[4:].strip()
            return json.loads(text)
        except Exception as e:
            logger.warning(f"Gemini API call failed or rate-limited: {e}")
            return None
    return None

def generate_fallback_metadata(keywords, pid):
    kw = str(keywords).lower()
    
    # 1. Determine Category
    category = "casual"
    if any(k in kw for k in ["saree", "kurta", "sherwani", "ethnic", "nehru", "chaniya", "lehenga", "anarkali"]):
        category = "ethnic"
    elif any(k in kw for k in ["gown", "dress", " farewell", " blazer", " tuxedo"]):
        category = "western"
    elif any(k in kw for k in ["jacket", "sweater", "fleece", "hoodie", "cardigan", "muffler", "shawl"]):
        category = "winter"
    elif any(k in kw for k in ["jogger", "hoodie", "sneaker", "streetwear"]):
        category = "streetwear"
    
    # 2. Determine Material
    material = "Cotton"
    for mat in ["silk", "linen", "velvet", "denim", "wool", "georgette", "chiffon", "crepe", "satin", "leather", "fleece", "suede"]:
        if mat in kw:
            material = mat.capitalize()
            break
            
    # 3. Determine Color
    color = "Multi"
    for col in ["red", "yellow", "maroon", "black", "white", "gold", "blue", "green", "pink", "grey", "brown", "beige", "cream", "orange"]:
        if col in kw:
            color = col.capitalize()
            break
            
    # 4. Determine Nature
    nature = "casual"
    if any(k in kw for k in ["puja", "diwali", "chhath", "bihar diwas", "wedding", "sangeet", "festive", "ceremonial", "haldi"]):
        nature = "festive"
    elif any(k in kw for k in ["farewell", "graduation", "formal", "office", "interview"]):
        nature = "formal"
        
    # 5. Age Range
    age_range = "Gen Z"
    if any(k in kw for k in ["wedding", "saree", "formal", "office", "farewell", "classic", "elegant"]):
        age_range = "Millennial"
        
    # 6. Generate Contextual Description (no hashtags)
    desc = f"An elegant {color.lower()} {material.lower()} garment designed beautifully for any occasion. Featuring a premium fit and comfort details."
    if "saree" in kw:
        desc = f"A gorgeous {color.lower()} {material.lower()} saree designed with beautiful traditional details. Drapes elegantly and offers a lightweight comfort, making it a perfect choice for festive celebrations."
    elif "kurta" in kw:
        desc = f"A classic {color.lower()} {material.lower()} kurta set designed for a smart look. Features a premium texture and lightweight weave that is comfortable to wear for celebrations and ceremonies."
    elif "jacket" in kw or "hoodie" in kw or "sweater" in kw:
        desc = f"A stylish and warm {color.lower()} {material.lower()} jacket. Designed with premium insulation and modern fit, making it perfect for winter layering."
    elif "gown" in kw or "dress" in kw:
        desc = f"A sophisticated {color.lower()} {material.lower()} dress featuring a sleek cut and premium drape. Perfect for farewell parties, dinners, or formal evening gatherings."
    else:
        # Custom build description from keyword chunks
        clean_kw = keywords.replace("", "'").replace("'", "")
        desc = f"A premium quality {clean_kw.lower()}. Crafted from fine {material.lower()} fabric, offering both comfort and style, perfectly suited for {nature} styling."

    return {
        "contextual_description": desc,
        "material": material,
        "color": color,
        "nature": nature,
        "age_range": age_range,
        "category": category
    }

def process_batch2(limit=None):
    if not os.path.exists(EXCEL_FILE):
        logger.error(f"Excel file not found at {EXCEL_FILE}")
        return

    # Load local_catalog.json
    with open(LOCAL_CATALOG_FILE, "r", encoding="utf-8") as f:
        catalog = json.load(f)
    catalog_map = {p["id"]: p for p in catalog}

    # Open Excel sheet
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active

    logger.info(f"Loaded Excel rows. Total items to check: {ws.max_row}")

    supabase = None
    if SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY:
        try:
            from supabase import create_client
            supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
            logger.info("Connected to Supabase.")
        except Exception as e:
            logger.error(f"Failed to connect to Supabase: {e}")

    processed_count = 0
    
    # We iterate K from 1 to 50
    for k in range(1, 51):
        row_idx = 2 * k - 1
        row = ws[row_idx]
        
        keywords = row[0].value
        url_a = row[1].value
        url_b = row[2].value
        
        if not keywords:
            logger.warning(f"Empty keywords at row {row_idx}, skipping.")
            continue
            
        # Prices and inventories from fallback rules or dynamic seeding
        # Let's seed prices dynamically based on K or a set ranges, matching Fashion Apparel 1
        price_a = int((k * 37) % 2500 + 799)
        price_b = int((k * 43) % 2500 + 799)
        
        inv_a = int((k * 13) % 40 + 5)
        inv_b = int((k * 17) % 40 + 5)

        items_to_process = [
            {"pid": 100 + (2 * k - 1), "url": url_a, "price": price_a, "inventory": inv_a, "suffix": "a"},
            {"pid": 100 + (2 * k), "url": url_b, "price": price_b, "inventory": inv_b, "suffix": "b"}
        ]

        for item in items_to_process:
            pid = item["pid"]
            url = item["url"]
            price = item["price"]
            inventory = item["inventory"]

            if not url:
                logger.warning(f"Empty URL for Product ID {pid}, skipping.")
                continue

            # Ensure catalog structure has placeholder entry if it doesn't exist
            if pid not in catalog_map:
                catalog_map[pid] = {
                    "id": pid,
                    "name": " ".join(keywords.split()[:5]).strip().capitalize(),
                    "description": keywords,
                    "category": "casual",
                    "image_url": f"/images/{pid}.webp",
                    "product_url": url,
                    "tags": [],
                    "zip_codes": []
                }
                # Add to catalog list
                catalog.append(catalog_map[pid])

            p_data = catalog_map[pid]

            # Resume support: skip if already has material and color
            if p_data.get("material") and p_data.get("color") and not limit:
                logger.info(f"Product {pid} already has enriched metadata, skipping.")
                continue

            # Find matching image path in destination public folder
            image_path = find_image_file(pid)
            if not image_path:
                logger.warning(f"Image for Product ID {pid} not found in public images directory!")
                # use fallback fallback
                image_path = find_image_file(1)

            logger.info(f"Processing Product ID {pid} ({keywords})...")
            fallback_used = False
            llm_data = None
            if image_path:
                llm_data = call_gemini_vision(image_path, keywords)

            if not llm_data:
                logger.warning(f"Failed to get LLM response for ID {pid}, using fallback generator...")
                llm_data = generate_fallback_metadata(keywords, pid)
                fallback_used = True

            # Update catalog item
            p_data["name"] = " ".join(keywords.split()[:5]).strip().capitalize()
            p_data["description"] = llm_data["contextual_description"]
            p_data["price"] = price
            p_data["inventory"] = inventory
            p_data["material"] = llm_data["material"]
            p_data["color"] = llm_data["color"]
            p_data["nature"] = llm_data["nature"]
            p_data["age_range"] = llm_data["age_range"]
            p_data["category"] = llm_data["category"].lower()
            p_data["product_url"] = url

            # Set correct image url extension based on what was copied
            if image_path:
                p_data["image_url"] = f"/images/{os.path.basename(image_path)}"

            # Vector generation
            import numpy as np
            new_tags = set(p_data.get("tags", []))
            new_tags.add(llm_data["material"].lower())
            new_tags.add(llm_data["color"].lower())
            new_tags.add(llm_data["nature"].lower())
            new_tags.add(llm_data["age_range"].lower())
            new_tags.add(llm_data["category"].lower())
            p_data["tags"] = list(new_tags)

            # Generate vibe vector
            vec = np.zeros(512)
            for t in p_data["tags"]:
                seed = sum(ord(c) for c in t)
                np.random.seed(seed)
                vec += np.random.randn(512) * 0.1
            norm = np.linalg.norm(vec)
            if norm > 0:
                vec = vec / norm
            p_data["embedding"] = vec.tolist()

            # Sync to Supabase
            if supabase:
                try:
                    payload = {
                        "id": pid,
                        "name": p_data["name"],
                        "description": p_data["description"],
                        "image_url": p_data["image_url"],
                        "product_url": p_data["product_url"],
                        "tags": p_data["tags"],
                        "zip_codes": p_data.get("zip_codes", []),
                        "price": price,
                        "inventory": inventory,
                        "material": p_data["material"],
                        "color": p_data["color"],
                        "nature": p_data["nature"],
                        "age_range": p_data["age_range"],
                        "category": p_data.get("category", "casual")
                    }
                    supabase.table("products").upsert(payload).execute()
                    logger.info(f"  Upserted/Synced Product ID {pid} to Supabase successfully.")
                except Exception as e:
                    logger.error(f"  Failed to sync Product ID {pid} to Supabase: {e}")

            processed_count += 1

            # Save local catalog incrementally
            with open(LOCAL_CATALOG_FILE, "w", encoding="utf-8") as f:
                json.dump(catalog, f, ensure_ascii=False, indent=2)

            if not fallback_used:
                time.sleep(13.5)
            else:
                time.sleep(0.1)

            if limit and processed_count >= limit:
                logger.info(f"Limit of {limit} reached, stopping.")
                break
        if limit and processed_count >= limit:
            break

    # Re-write frontend fallback file as well
    FRONTEND_FALLBACK_FILE = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "frontend", "src", "catalog_fallback.js"))
    if os.path.exists(FRONTEND_FALLBACK_FILE):
        try:
            js_content = "export const FALLBACK_PRODUCTS = " + json.dumps(catalog, indent=2) + ";\n"
            with open(FRONTEND_FALLBACK_FILE, "w", encoding="utf-8") as f:
                f.write(js_content)
            logger.info("Rewrote frontend catalog fallback file.")
        except Exception as e:
            logger.error(f"Failed to rewrite frontend fallback: {e}")

    logger.info(f"Vision Enrichment for Batch 2 finished. Processed {processed_count} items.")

if __name__ == "__main__":
    limit_val = None
    if len(sys.argv) > 1:
        try:
            limit_val = int(sys.argv[1])
        except ValueError:
            pass
    process_batch2(limit=limit_val)
