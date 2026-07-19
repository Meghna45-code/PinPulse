import openpyxl

def inspect_file(filepath):
    print(f"\n=== Inspecting {filepath} ===")
    wb = openpyxl.load_workbook(filepath, data_only=True) # data_only=True to get raw values, not formulas
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"Sheet: {name}, Rows: {ws.max_row}, Cols: {ws.max_column}")
        for i in range(1, min(6, ws.max_row + 1)):
            row_vals = [cell.value for cell in ws[i]]
            print(f"  Row {i}: {row_vals}")

inspect_file("Fashion Apparel.xlsx")
inspect_file("Fashion Apparel2.xlsx")
